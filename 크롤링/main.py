from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook

baseUrl = r"https://sanjaecase.comwel.or.kr/service/dataView?id="
nameUrl = r"https://sanjaecase.comwel.or.kr/service/dataListReverse?gubun2=%EC%9E%91%EC%97%85%EC%8B%9C%EA%B0%84%EC%A4%91%EC%82%AC%EA%B3%A0&gubun=&pageUnit=200&pageIndex="

driver = webdriver.Chrome()
driver2 = webdriver.Chrome()
wb = load_workbook(filename="크롤링/판례.xlsx")
ws = wb.active

for i in range(46):  # 46
    driver.get(nameUrl + str(i+1))
    driver.maximize_window()
    WebDriverWait(driver, 30).until(EC.presence_of_element_located(
        (By.XPATH, f'/html/body/div/div[2]/div[2]/div/div[3]/div[2]/div[1]/table/tbody')))

    for j in range(200):  # 200
        print(f"{i+1}페이지 {j+1}번째")
        num = i * 200 + j + 1
        casetype = driver.find_element(
            By.XPATH, f"/html/body/div/div[2]/div[2]/div/div[3]/div[2]/div[1]/table/tbody/tr[{j+1}]/td[2]").text
        casenum = driver.find_element(
            By.XPATH, f"/html/body/div/div[2]/div[2]/div/div[3]/div[2]/div[1]/table/tbody/tr[{j+1}]/td[3]/a/span").text[1:-1]
        casenum += "_"
        casenum += driver.find_element(
            By.XPATH, f"/html/body/div/div[2]/div[2]/div/div[3]/div[2]/div[1]/table/tbody/tr[{j+1}]/td[3]/a").text.split("\n")[0]
        casename = driver.find_element(
            By.XPATH, f"/html/body/div/div[2]/div[2]/div/div[3]/div[2]/div[1]/table/tbody/tr[{j+1}]/td[4]").text

        ws.cell(row=num+1, column=1).value = num
        ws.cell(row=num+1, column=2).value = casetype
        ws.cell(row=num+1, column=3).value = casenum
        ws.cell(row=num+1, column=4).value = casename

        driver2.get(baseUrl + casenum)
        WebDriverWait(driver2, 30).until(
            EC.presence_of_element_located((By.XPATH, "/html/body/div/div[2]")))

        pl = driver2.find_element(
            By.XPATH, "/html/body/div/div[2]/div[2]/div[1]/ul[2]/li[2]")

        ddt = driver2.find_element(
            By.XPATH, "/html/body/div/div[2]/div[2]/div[1]/ul[3]/li[2]")
        try:
            sentence = driver2.find_element(
                By.XPATH, "/html/body/div/div[2]/div[2]/div[1]/ul[6]/li[2]").text

            close = driver2.find_element(
                By.XPATH, "/html/body/div/div[2]/div[2]/div[1]/ul[5]/li[2]").text
        except:
            try:
                sentence = driver2.find_element(
                    By.XPATH, "/html/body/div/div[2]/div[2]/div[1]/ul[5]/li[2]").text
                close = driver2.find_element(
                    By.XPATH, "/html/body/div/div[2]/div[2]/div[1]/ul[4]/li[2]").text
            except:
                sentence = "변론종결되지 않음"
                close = "판결선고되지 않음"

        ws.cell(
            row=num+1, column=5).value = " | ".join(x for x in pl.text.split("\n") if x != "\n")
        ws.cell(row=num+1, column=6).value = " | ".join(
            x for x in ddt.text.split("\n") if x != "\n")
        ws.cell(row=num+1, column=7).value = close
        ws.cell(row=num+1, column=8).value = sentence

wb.save("result.xlsx")
driver.quit()
driver2.quit()
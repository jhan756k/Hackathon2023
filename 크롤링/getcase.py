from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
import win32api, time

baseUrl = r"https://sanjaecase.comwel.or.kr/service/dataView?id="

driver = webdriver.Chrome()

wb = load_workbook(filename=r"C:\Users\Jooney Han\Desktop\해커톤\크롤링\result.xlsx")
ws = wb.active

for i in range(9200): # 9200
    print(f"{i+1}번째")
    case = ws.cell(row=i+2, column=3).value
    driver.get(baseUrl + case)
    
    try:
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.XPATH, "/html/body/div/div[2]")))
    except:
        driver.refresh()
        win32api.MessageBox(0, 'hello', 'title', 0x00001000)
        time.sleep(5)
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.XPATH, "/html/body/div/div[2]")))
    
    fullcase = driver.find_element(By.XPATH, "/html/body/div/div[2]/div[2]/div[2]").text
    ws.cell(row=i+2, column=9).value = fullcase

wb.save("result.xlsx")
driver.quit()
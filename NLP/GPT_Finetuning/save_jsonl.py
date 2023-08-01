from openpyxl import load_workbook
import re

f = open("data.jsonl", "w", encoding="utf-8")

wb = load_workbook(filename=r"C:\Users\Jooney Han\Desktop\해커톤\크롤링\result.xlsx")
ws = wb.active

for i in range(10): # 9062 - 판결있는거만
    celldata = ws.cell(row=i+2, column=9).value
    
    casename = ws.cell(row=i+2, column=4).value
    casename += ", "
    casename += ws.cell(row=i+2, column=2).value

    r1 = celldata.find("이유\n")
    r2 = celldata.find("결론\n")
    reason = celldata[r1:r2]
    
    prompt = "사건명: " + casename + "\n이유: " + reason
    
    c1 = celldata.find("결론\n")
    completion = celldata[c1:]

    prompt, completion = prompt.replace('"',r'\"'), completion.replace('"',r'\"')
    f.write(f"{{\"prompt\": \"{prompt}\", \"completion\": \"{completion}\"}}\n")

f.close()